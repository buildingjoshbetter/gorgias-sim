#!/usr/bin/env python3
"""P2 — per-county fiscal impact of raising the residence homestead exemption.

Fetches the Comptroller 2025 ISD rates & levies XLSX (the residence homestead
exemption applies to *school district* M&O taxes, so the school levy is the honest
basis), aggregates school-district M&O levy per county, and computes a per-county
impact of the hero scenario (raise the general residence homestead exemption from
$100,000 to $140,000) using the Appendix A.6 model.

The number is a MODEL on real data, not an official estimate. Exact arithmetic is
recorded in the `methodology` string that ships in the JSON.

Impact model (Appendix A.6 — "school-district levy share attributable to the
homestead base x delta ratio"):

    school_mo_levy   = sum over ISDs in the county of (TAXABLE VALUE M&O x M&O RATE)
    homestead_share  = 0.45   # assumed statewide residence-homestead fraction of the
                              #   school M&O taxable base (documented constant)
    delta_ratio      = 40000 / 300000 = 0.1333   # +$40k exemption removed from an
                              #   assumed $300k average homestead taxable value
    impact_usd       = school_mo_levy * homestead_share * delta_ratio

This yields the annual reduction in local school M&O revenue per county. The map and
top-10 ordering are therefore driven by each county's *real* school M&O levy; the two
modeling constants scale magnitude uniformly and are stated up front.
"""

import json
import os
import re
from datetime import datetime, timezone

import httpx
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.abspath(os.path.join(HERE, "..", "data"))
GEOJSON = os.path.join(DATA, "tx-counties.svg.json")
OUT = os.path.join(DATA, "county-impact.json")
CACHE = os.path.join(HERE, "out", "raw")

DOCS = "https://comptroller.texas.gov/taxes/property-tax/docs/"
ISD_URL = DOCS + "2025-school-district-rates-levies.xlsx"

# All 2025 taxing-unit levy workbooks — summed for the total-levy denominator.
ALL_UNIT_FILES = [
    "2025-county-rates-levies.xlsx",           # county general/road & bridge funds
    "2025-school-district-rates-levies.xlsx",  # ISDs
    "2025-city-rates-levies.xlsx",             # cities
    "2025-special-district-rates-levies.xlsx", # special districts (MUDs, hospital, etc.)
]

HOMESTEAD_SHARE = 0.45
# Baseline is the CURRENT $140,000 exemption (post-Prop 13, Nov 2025 / HB 9); the
# hero scenario raises it to $200,000, so the incremental exemption is $60,000.
DELTA_EXEMPTION = 60000.0
AVG_HOMESTEAD_TAXABLE = 300000.0
DELTA_RATIO = DELTA_EXEMPTION / AVG_HOMESTEAD_TAXABLE

METHODOLOGY = (
    "NUMERATOR (impact_usd) — modeled from Comptroller 2025 ISD rates & levies. "
    "Per county: school_mo_levy = sum over ISDs in county of "
    "(TAXABLE VALUE M&O x M&O RATE / 100), rates quoted per $100 of taxable value. "
    "impact_usd = school_mo_levy x homestead_share(0.45) x delta_ratio(60000/300000 "
    "= 0.2), the annual local school M&O revenue reduction from raising the general "
    "residence homestead exemption from its CURRENT $140,000 (post-Prop 13 2025 / HB 9) "
    "to $200,000. The 0.45 homestead-base share and $300k average homestead taxable "
    "value are stated modeling constants; magnitude is driven by each county's real "
    "school M&O levy. "
    "DENOMINATOR (impact_pct) — total_all_units_levy = sum of CALCULATED LEVY across "
    "every taxing unit levying in the county in the 2025 Comptroller rates & levies "
    "workbooks (county general/road & bridge funds + ISDs + cities + special "
    "districts). impact_pct = impact_usd / total_all_units_levy x 100, so it varies by "
    "county: school-tax-heavy counties score higher, counties with large city/special- "
    "district bases score lower. This is honest arithmetic on real data, not an "
    "official fiscal estimate."
)


def norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def fetch_file(fname):
    """Download a Comptroller levy workbook into the raw cache; return its path."""
    os.makedirs(CACHE, exist_ok=True)
    path = os.path.join(CACHE, fname)
    with httpx.Client(timeout=90.0, follow_redirects=True) as c:
        r = c.get(DOCS + fname)
        r.raise_for_status()
        with open(path, "wb") as fh:
            fh.write(r.content)
    return path


def header_index(ws):
    """Locate the header row (contains 'CAD ID') and return {col_name: idx}."""
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "CAD ID":
            return {str(v).strip(): i for i, v in enumerate(row) if v is not None}
    raise ValueError("no header row with 'CAD ID' found")


def aggregate_total_levy(xlsx_path):
    """Sum CALCULATED LEVY by COUNTY NAME across every taxing unit in a workbook."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = header_index(ws)
    ci, li = hdr["COUNTY NAME"], hdr["CALCULATED LEVY"]
    by_county = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] in (None, "", "CAD ID"):
            continue
        county = row[ci]
        levy = row[li]
        if county in (None, "") or levy in (None, ""):
            continue
        try:
            by_county[str(county).strip()] = (
                by_county.get(str(county).strip(), 0.0) + float(levy)
            )
        except (TypeError, ValueError):
            continue
    return by_county


def aggregate_school_levy(xlsx_path):
    """Return {county_name: mo_levy} summed across all ISD county-splits."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Detail"]
    header = None
    by_county = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if row and row[0] == "CAD ID":
                header = {str(v).strip(): i for i, v in enumerate(row) if v is not None}
            continue
        if not row or row[header["CAD ID"]] in (None, ""):
            continue
        county = str(row[header["COUNTY NAME"]]).strip()
        tv_mo = row[header["TAXABLE VALUE M&O"]]
        mo_rate = row[header["M&O RATE"]]
        if tv_mo in (None, "") or mo_rate in (None, ""):
            continue
        try:
            # Texas tax rates are quoted per $100 of taxable value.
            mo_levy = float(tv_mo) * float(mo_rate) / 100.0
        except (TypeError, ValueError):
            continue
        by_county[county] = by_county.get(county, 0.0) + mo_levy
    return by_county


def main():
    counties = json.load(open(GEOJSON))  # [{fips, name, d}] — 254 rows

    # Numerator basis: school M&O levy per county (unchanged).
    school_levy = aggregate_school_levy(fetch_file("2025-school-district-rates-levies.xlsx"))
    levy_idx = {norm(k): (k, v) for k, v in school_levy.items()}

    # Denominator basis: total CALCULATED LEVY across ALL taxing units, all workbooks.
    total_levy = {}
    for fname in ALL_UNIT_FILES:
        for county, lv in aggregate_total_levy(fetch_file(fname)).items():
            total_levy[county] = total_levy.get(county, 0.0) + lv
    total_idx = {norm(k): v for k, v in total_levy.items()}

    rows = []
    gaps = []
    for c in counties:
        fips = c["fips"]
        name = c["name"]
        hit = levy_idx.get(norm(name))
        if hit is None:
            rows.append(
                {"fips": fips, "name": name, "impact_usd": None,
                 "impact_pct": None, "reason": "no matching county in ISD levy file"}
            )
            gaps.append(name)
            continue
        mo_levy = hit[1]
        impact_usd = mo_levy * HOMESTEAD_SHARE * DELTA_RATIO
        total = total_idx.get(norm(name))
        if total and total > 0:
            # Percentage of the county's entire property-tax levy (all taxing units).
            impact_pct = round(impact_usd / total * 100.0, 2)
        else:
            impact_pct = None
        rows.append(
            {"fips": fips, "name": name,
             "impact_usd": round(impact_usd, 2),
             "impact_pct": impact_pct}
        )

    rows.sort(key=lambda r: r["fips"])
    out = {
        "counties": rows,
        "methodology": METHODOLOGY,
        "source_url": ISD_URL,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
    }
    with open(OUT, "w") as fh:
        json.dump(out, fh, separators=(",", ":"))

    valued = [r for r in rows if r["impact_usd"] is not None]
    top_usd = sorted(valued, key=lambda r: r["impact_usd"], reverse=True)[:10]
    top_pct = sorted([r for r in valued if r["impact_pct"] is not None],
                     key=lambda r: r["impact_pct"], reverse=True)[:10]
    print(f"county rows: {len(rows)}  (with impact: {len(valued)}, gaps: {len(gaps)})")
    if gaps:
        print("DATA GAPS:", ", ".join(gaps))
    print(f"file size: {os.path.getsize(OUT)} bytes")
    print("\nTOP 10 BY impact_usd:")
    for i, r in enumerate(top_usd, 1):
        print(f"  {i:2d}. {r['name']:<14} FIPS {r['fips']}  "
              f"${r['impact_usd']:,.0f}  ({r['impact_pct']}% of total levy)")
    print("\nTOP 10 BY impact_pct (school-tax-heavy counties):")
    for i, r in enumerate(top_pct, 1):
        print(f"  {i:2d}. {r['name']:<14} FIPS {r['fips']}  "
              f"{r['impact_pct']}%  (${r['impact_usd']:,.0f})")


if __name__ == "__main__":
    main()
